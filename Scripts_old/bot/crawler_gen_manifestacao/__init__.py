from Scripts.bot.jobs.append_info import replace_placeholder
from Scripts.bot.jobs.croppdf import crop_itau
from Scripts.bot.jobs.copy_files import copy
from Scripts.bot.jobs.read_pdf import get_barcode
from Scripts.bot.jobs.makeoutput import MakeXlsx

import platform
import subprocess
import glob
import pathlib
import os
import openpyxl
import time
from typing import Type
from termcolor import colored
from openpyxl.worksheet.worksheet import Worksheet
import locale
from datetime import datetime
from pytz import timezone
from Scripts.Tools.PrintLogs import printtext as prt
from Scripts.Tools.StartStop_Notify import SetStatus


        

class GeradordeDoc:
    
    def __init__(self, arguments_bot: dict) -> None:
        
        self.soffice = None
        self.set_status = SetStatus()
        self.arguments_bot = arguments_bot
        self.pid = str(arguments_bot["pid"])
        self.prt = prt(self.pid)
        
        self.prt.print_log('log', 'Robô inicializado!')
        
        self.input_file = glob.glob(os.path.join("Temp", self.pid, "*.xlsx"))[0]
        
        status = [self.arguments_bot['user'], self.pid, "crawler_gen_manifestacao", 'Em Execução', self.input_file]
        self.set_status.botstart(status)
        
        self.pathfiles = pathlib.Path(self.input_file).parent.resolve().__str__()
        
        self.path_output = pathlib.Path(self.input_file).parent.resolve().__str__()
        self.ws: Type[Worksheet] = openpyxl.load_workbook(self.input_file).active
        self.row = 2
        self.index = 0
        os.makedirs(os.path.join(self.path_output, "protocolar"), exist_ok=True)
        
        self.path = os.path.join(self.path_output, f"{self.pid} - Peticionamento {datetime.now(timezone('Etc/GMT+4')).strftime('%d-%m-%y')}.xlsx")
        MakeXlsx().make_output("peticionamento", self.path)
        self.codbarsws = os.path.join(self.path_output, f"{self.pid} - Código de Barras {datetime.now(timezone('Etc/GMT+4')).strftime('%d-%m-%y')}.xlsx")
        MakeXlsx().make_output(f"cod_bars", self.codbarsws)
        self.start_time = time.perf_counter()
        
    def execution(self):
        
        self.bot_data = {}
        self.prt = prt(self.pid, self.row-1)
        if self.row == self.ws.max_row+1:
            end_time = time.perf_counter()
            execution_time = end_time - self.start_time
            calc = execution_time/60
            splitcalc = str(calc).split(".")
            minutes = int(splitcalc[0])
            seconds = int(float(f"0.{splitcalc[1]}") * 60)
            self.prt.print_log("log", f"Fim da execução, tempo: {minutes} minutos e {seconds} segundos")
            
            barra = "\\" if "\\" in self.path else "/"
            
            namefile = self.path.split(barra)[-1]
            status = [self.arguments_bot['user'], self.pid, self.arguments_bot.get('bot'), 'Finalizado', namefile]
            self.set_status.botstop(status)
            return
            
        for index in range(1, self.ws.max_column+1):
            self.index = index
            self.bot_data.update(self.set_data())
            if index  == self.ws.max_column:
                break
        
        
        if len(self.bot_data) == 0:
            self.row +=1
            self.execution()
        
        self.queue()
        
    def queue(self):
        
        try:
            
            data_concatenada = str(self.bot_data.get("[[CIDADE_ESTADO]]") + ", " + self.bot_data.get("[[DATA]]"))
            self.bot_data.update({"[[DATA]]": data_concatenada})
            self.bot_data.pop("[[CIDADE_ESTADO]]")
            files_peticao = copy(self.bot_data.get("[[NUMERO_PROCESSO]]"), self.pathfiles, self.codbarsws, self.pid)
            if not len(files_peticao) == 0:
                img = crop_itau(files_peticao[0])
                self.bot_data.setdefault("[[COMPROVANTE]]", img)
                name_peticao = replace_placeholder(self.path_output, self.bot_data, self.pid)
                
                docs_anexos = None
                for anexos in files_peticao:
                    anexos = str(anexos)
                    barra = "\\" if "\\" in anexos else "/"
                    if docs_anexos is None:
                        docs_anexos = anexos.split(barra)[-1]
                        continue
                    docs_anexos = docs_anexos + "," + anexos.split(barra)[-1]
                
                name_peticao = name_peticao.split(barra)[-1]
                
                self.convert_to_pdf(os.path.join(self.path_output, "protocolar", name_peticao))
                self.convert_to_pdf(os.path.join(self.path_output, self.bot_data.get("[[NUMERO_PROCESSO]]"), name_peticao))
                
                data = [self.bot_data.get("[[NUMERO_PROCESSO]]"), "PETIÇÃO - MANIFESTAÇÃO DA PARTE", "Petição", name_peticao.replace(".docx", ".pdf"), docs_anexos, "Petição"]
                self.append_success(data)
                
            self.row +=1
            self.execution()
            
        
        except Exception as e:
            
            self.prt.print_log("error", f"{e}")
            self.row +=1
            self.execution()
            
            
        
    def set_data(self) -> dict:

        returns = {}
        for nome_coluna in self.nomes_colunas():
            nome_coluna = str(nome_coluna)
            nome_coluna_planilha = self.ws.cell(row=1, column=self.index).value
            valor_celula = self.ws.cell(row=self.row, column=self.index).value
            if nome_coluna_planilha and nome_coluna.upper() == str(nome_coluna_planilha).upper():
                if valor_celula:
                    
                    if isinstance(valor_celula, datetime):
                        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
                        valor_celula = valor_celula.strftime('%d de %B de %Y')
                    
                    returns = {f'[[{nome_coluna.upper()}]]': valor_celula}
                    break
            
        return returns
    
    def nomes_colunas(self) -> list:
    
        all_fields = [
            "NUMERO_PROCESSO", "COMARCA", "DATA", "CIDADE_ESTADO"
        ]
        return all_fields
    
    def append_success(self, data: list, message: str = None):

        wb = openpyxl.load_workbook(filename=self.path)
        sheet = wb.active

        sheet.append(data)
        wb.save(self.path)
        
        if not message:
            self.message = f'Execução do processo Nº{data[0]} efetuada com sucesso!'
        
        self.prt.print_log("log", self.message)

    def convert_to_pdf(self, input_path: str):

        path = 'c:\\Program Files\\'
        name = 'soffice.exe'
        
        currentdir = pathlib.Path(input_path).parent.resolve().__str__()
        arquivo_de_entrada = input_path
        outfile = os.path.join(currentdir)
        if platform.system() == 'Linux':
            try:
                subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', outfile , arquivo_de_entrada], check=True)
            except Exception as e:
                print(e)
        else:
            if not self.soffice:
                for root, dirs, files in os.walk(path):
                    if name in files:
                        self.soffice = os.path.join(root, name)
            
            try:
                subprocess.run([self.soffice, '--headless', '--convert-to', 'pdf', '--outdir', outfile , arquivo_de_entrada], check=True)
            except Exception as e:
                print(e)