import os
import time
import shutil
import pathlib
import openpyxl
import unicodedata
from time import sleep
from typing import Type
from contextlib import suppress


""" Imports do Projeto """
from bot.head import CrawJUD

from bot.head.Tools.PrintLogs import printtext as prt
from bot.head.common.exceptions import ErroDeExecucao
from bot.head.common.selenium_excepts import webdriver_exepts
from bot.head.common.selenium_excepts import exeption_message


# Selenium Imports
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import  NoSuchElementException, TimeoutException


class protocolo(CrawJUD):

    def __init__(self, Initbot: Type[CrawJUD]) -> None:
        
        self.__dict__ = Initbot.__dict__.copy()
        self.start_time = time.perf_counter()
        
    def execution(self):
        
        while not self.thread._is_stopped:
            if self.row == self.ws.max_row+1:
                self.prt = prt(self.pid, self.row)
                break
            
            self.bot_data = {}
            for index in range(1, self.ws.max_column + 1):
                self.index = index
                self.bot_data.update(self.set_data())
                if index == self.ws.max_column:
                    break
            
            try:
                
                if not len(self.bot_data) == 0:
                    self.prt = prt(self.pid, self.row-1, url_socket=self.argbot['url_socket'])
                    self.queue()
                
            except Exception as e:
                
                old_message = self.message
                self.message = getattr(e, 'msg', getattr(e, 'message', ""))
                if self.message == "":
                    for exept in webdriver_exepts():
                        if isinstance(e, exept):
                            self.message = exeption_message().get(exept)
                            break
                        
                if not self.message:
                    self.message = str(e)
                    
                error_message = f'{self.message}. | Operação: {old_message}'
                self.prt.print_log("error", error_message)
                self.append_error([self.bot_data.get('NUMERO_PROCESSO'), self.message])
            
            self.row += 1
            
        self.finalize_execution()

        
    def queue(self):

        self.search(self.bot_data, self.prt)
        self.init_protocolo()
        self.set_tipo_protocolo()
        self.set_subtipo_protocolo()
        self.set_petition_file()
        self.vincular_parte()
        self.finish_petition()
        data = self.get_confirm_protocol()
        self.append_success(data, message = data[1])

    def init_protocolo(self):
        
        try:
            try:
                self.prt.print_log('log', 'Processo encontrado! Inicializando peticionamento...')
                button_peticionamento:WebElement = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID,'pbPeticionar')))
                link = button_peticionamento.get_attribute('onclick').split("'")[1]
                self.driver.execute_script("return window.location.href = '{link}';".format(link=link))
                sleep(5)

            except:

                button_enterproc:WebElement = WebDriverWait(self.driver, 5).until(EC. presence_of_element_located((By.CSS_SELECTOR, '#processoSelecionado')))
                button_enterproc.click()

                enterproc:WebElement = WebDriverWait(self.driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#botaoEnviarIncidente')))
                enterproc.click()
                button_peticionamento:WebElement = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID,'pbPeticionar')))
                link = button_peticionamento.get_attribute('onclick').split("'")[1]
                self.driver.execute_script("return window.location.href = '{link}';".format(link=link))
                
        except:
            raise ErroDeExecucao("Erro ao inicializar peticionamento") 
    
    def set_tipo_protocolo(self):

        try:
            self.interact.sleep_load('div[id="loadFeedback"]')
            self.prt.print_log('log', 'Informando tipo de peticionamento')
            button_classification: WebElement = self.wait.until(EC.presence_of_element_located((By.ID,'botaoEditarClassificacao')))
            self.interact.click(button_classification)
            
            select_tipo_peticao: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.ui-select-container[input-id="selectClasseIntermediaria"]')))
            select_tipo_peticao = select_tipo_peticao.find_element(By.CSS_SELECTOR, 'span.btn.btn-default.form-control.ui-select-toggle')
            self.interact.click(select_tipo_peticao)
            
            input_tipo_peticao: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input#selectClasseIntermediaria')))
            self.interact.send_key(input_tipo_peticao, self.bot_data.get('TIPO_PROTOCOLO'))
            sleep(1.5)
            self.interact.send_key(input_tipo_peticao, Keys.ENTER)
        
        except:
            raise ErroDeExecucao("Erro ao informar tipo de protocolo") 
    
    def set_subtipo_protocolo(self):
        
        try:
            self.prt.print_log('log', 'Informando subtipo de peticionamento')
            select_categoria_peticao: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.ui-select-container[input-id="selectCategoria"]')))
            select_categoria_peticao = select_categoria_peticao.find_element(By.CSS_SELECTOR, 'span.btn.btn-default.form-control.ui-select-toggle')
            self.interact.click(select_categoria_peticao)

            input_categoria_peticao: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input#selectCategoria')))
            self.interact.send_key(input_categoria_peticao, self.bot_data.get("SUBTIPO_PROTOCOLO"))
            
            input_categoria_peticao_option: WebElement = self.wait.until(EC.presence_of_element_located((By.XPATH,'.//li[@class="ui-select-choices-group"]/ul/li/span')))
            input_categoria_peticao_option.click()
            sleep(1)
        
        except:
            raise ErroDeExecucao("Erro ao informar subtipo de protocolo") 
    
    def set_petition_file(self):
        
        try:
            self.prt.print_log('log', 'Anexando petição')
            input_file: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#botaoAdicionarDocumento > input[type=file]')))
            sleep(2)
            
            path_file = pathlib.Path(self.input_file).parent.resolve().__str__()
            file = os.path.join(path_file, self.bot_data.get('PETICAO_PRINCIPAL'))
            
            file = file.replace(" ", "")
            if "_" in file:
                file = file.replace("_","")
            
            
            file = unicodedata.normalize('NFKD', file)
            file = "".join([c for c in file if not unicodedata.combining(c)])
            
            input_file.send_keys(file)
            
            file_uploaded = ""
            with suppress(TimeoutException):
                file_uploaded:WebElement = WebDriverWait(self.driver, 25).until(EC.presence_of_element_located((By.XPATH,'//nav[@class="document-data__nav"]/div/ul/li[5]/button[2]')))
            
            if file_uploaded == "":
                raise ErroDeExecucao("Erro ao enviar petição") 
                
            self.prt.print_log('log', 'Petição do processo anexada com sucesso')
        
        except:
            raise ErroDeExecucao("Erro ao enviar petição") 
    
    def vincular_parte(self):
        
        try:
            parte_peticao = self.bot_data.get("PARTE_PETICIONANTE").__str__().lower()
            self.prt.print_log('log', 'Vinculando parte a petição...')
            partes: WebElement = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[ui-view="parteProcessoView"]')))
            if partes:
                for parte in partes:
                    parte: WebElement = parte
                    parte_name = parte.find_element(By.CSS_SELECTOR, 'span[ng-bind="parte.nome"]').text.lower()
                    if parte_name == parte_peticao:
                        
                        sleep(3)
                        
                        incluir_button = None
                        with suppress(NoSuchElementException):
                            incluir_button = parte.find_element(By.CSS_SELECTOR, 'button[ng-click="incluirParteDoProcessoPeticaoDiversa(parte)"]')
                            
                        if not incluir_button:
                            with suppress(NoSuchElementException):
                                incluir_button = parte.find_element(By.CSS_SELECTOR, 'button[ng-click="incluirParteDoProcessoNoPoloContrario(parte)"]')
                                
                        incluir_button.click()
                            
                        self.prt.print_log('log', 'Vinculando cliente à petição...')
                        sleep(0.3)
                        break   
                        
                    else:
                        partes = self.driver.find_elements(By.CSS_SELECTOR, 'div[ui-view="parteView"]')
                        for parte in partes:
                            parte_name = parte.find_element(By.CSS_SELECTOR, 'span[ng-bind="parte.nome"]').text.lower()
                            if parte_name == parte_peticao.lower():
                                self.prt.print_log('log', 'Parte já vinculada, finalizando peticionamento...')
                                sleep(0.3)
                                break

            else: 
                raise ErroDeExecucao("Não foi possivel vincular parte a petição") 
                
        except:
            raise ErroDeExecucao("Não foi possivel vincular parte a petição") 
    
    def finish_petition(self):
        
        self.prt.print_log('log', 'Finalizando...')

        finish_button = self.driver.find_element(By.XPATH, '//*[@id="botaoProtocolar"]')
        sleep(1)
        finish_button.click()
        sleep(5)

        confirm_button: WebElement = self.wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'div.popover-content button.confirm-button')))
        confirm_button.click()
    
    def get_confirm_protocol(self) -> list:
        
        try:
            getlinkrecibo:WebElement = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'button[ng-click="consultarReciboPeticao(peticao)"]')))
            
            sleep(3)
            
            name_recibo = f"Recibo Protocolo - {self.bot_data.get('NUMERO_PROCESSO')} - PID {self.pid}.pdf"
            self.driver.get_screenshot_as_file(f'{self.output_dir_path}/{name_recibo.replace(".pdf", ".png")}')
            
            getlinkrecibo.click()
            
            path = os.path.join(self.output_dir_path, name_recibo)
            pathpdf = os.path.join(pathlib.Path(self.input_file).parent.resolve(), 'recibo.pdf')
            
            while True:
                
                if os.path.exists(pathpdf):
                    sleep(0.5)
                    break
            
            shutil.move(pathpdf, path)
            return [self.bot_data.get('NUMERO_PROCESSO'), f"Processo nº{self.bot_data.get('NUMERO_PROCESSO')} protocolado com sucesso!", name_recibo]
        
        except Exception as e:
            raise ErroDeExecucao("Erro ao confirmar protocolo")
    