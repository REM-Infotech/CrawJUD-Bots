import os
import re
import time
import shutil
import pathlib
import openpyxl
import platform
import logging
import subprocess
import unicodedata
from time import sleep
from typing import Type
from contextlib import suppress


""" Imports do Projeto """
from bot.head import CrawJUD
from bot.head.count_doc import count_doc

from bot.head.Tools.PrintLogs import printtext as prt
from bot.head.common.exceptions import ErroDeExecucao
from bot.head.common.selenium_excepts import webdriver_exepts
from bot.head.common.selenium_excepts import exeption_message


# Selenium Imports
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import  NoSuchElementException, TimeoutException, StaleElementReferenceException
from PyPDF2 import *

class CaixaMaster(CrawJUD):

    def __init__(self, Initbot: Type[CrawJUD]) -> None:
        
        self.__dict__ = Initbot.__dict__.copy()
        
        self.start_time = time.perf_counter()
        
    def execution(self):
        
        while not self.thread._is_stopped:
            if self.row == self.ws.max_row+1:
                self.row = self.ws.max_row
                break
            
            self.bot_data = {}
            for index in range(1, self.ws.max_column + 1):
                self.index = index
                self.bot_data.update(self.set_data())
                if index == self.ws.max_column:
                    break
            
            try:
                
                if not len(self.bot_data) == 0:
                    self.queue()
                
            except Exception as e:
                
                old_message = self.message
                self.message = getattr(e, 'msg', getattr(e, 'message', ""))
                if self.message == "":
                    for exept in webdriver_exepts():
                        if isinstance(e, exept):
                            self.message = exeption_message().get(exept)
                            break
                        
                if not self.message:
                    self.message = str(e)
                
                self.type_log = "error"
                self.message_error = f'{self.message}. | Operação: {old_message}'
                self.prt(self)()
                self.append_error([self.bot_data.get('NUMERO_PROCESSO'), self.message])
                self.message_error = None
            
            self.row += 1
            
        self.finalize_execution()
        
    def queue(self) -> None:
        
        nameboleto = None
        self.get_site()
        self.locale_proc()
        self.proc_nattribut()
        self.dados_partes()
        self.info_deposito()
        self.make_doc()
        nameboleto = self.rename_pdf()
        data = self.get_val_doc_and_codebar(nameboleto)
        self.append_success(data)
                                
    def get_site(self) -> None:
        
        self.message = "Acessando página de emissão"
        self.type_log = "log"
        self.prt(self)()
        
        self.driver.get("https://depositojudicial.caixa.gov.br/sigsj_internet/depositos-judiciais/justica-estadual/")
        list_opt: WebElement = self.wait.until(EC.presence_of_element_located((By. CSS_SELECTOR, 'select[id="j_id5:filtroView:j_id6:tpDeposito"]')))
        list_options = list_opt.find_elements(By.TAG_NAME, 'option')

        for option in list_options:
            if option.text == "Depósitos Judiciais da Justiça Estadual":
                option.click()
                break
        
        captchainput: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[id="autoCaptcha"')))
        val_captcha = captchainput.get_attribute('value')
        
        inputcaptcha: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[id="j_id5:filtroView:j_id6:j_id17:captchaView:cpatchaTextBox"]')))
        inputcaptcha.send_keys(val_captcha.replace(",", ""))
        
        next_btn = self.driver.find_element(By.CSS_SELECTOR, 'input[class="hand btnConfirmar"]')
        next_btn.click()   
        
        sleep(2)
        next_btn: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'a[id="j_id5:filtroView:mensagemView:j_id77:btnProsseguir')))
        next_btn.click()
    
    def locale_proc(self) -> None:

        self.interact.wait_caixa()
        
        self.message = "Informando tribunal"
        self.type_log = "log"
        self.prt(self)()
        
        lista_tribunal: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:coTribunal"]'))).find_elements(By.TAG_NAME, 'option')
        for item in lista_tribunal:
            item: WebElement = item
            if str(self.bot_data.get("TRIBUNAL")).lower() in item.text.lower():
                item.click()
                break

        self.interact.wait_caixa() 
        
        self.message = "Informando comarca"
        self.type_log = "log"
        self.prt(self)()
        
        lista_comarca: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:coComarca"]'))).find_elements(By.TAG_NAME, 'option')
        for item in lista_comarca:
            item: Type[WebElement] = item
            if str(self.bot_data.get("COMARCA")).lower() in item.text.lower():
                item.click()
                break
            
        
        self.interact.wait_caixa() 
        self.message = "Informando vara"
        self.type_log = "log"
        self.prt(self)()
        lista_vara: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:coVara"]'))).find_elements(By.TAG_NAME, 'option')
        for item in lista_vara:
            item: WebElement = item
            if str(self.bot_data.get("VARA")).lower() in item.text.lower():
                item.click()
                break
                
        
        self.interact.wait_caixa()
        self.message = 'Informando agencia'
        self.type_log = "log"
        self.prt(self)()
        lista_agencia: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:coAgencia"]'))).find_elements(By.TAG_NAME, 'option')
        for item in lista_agencia:
            item: WebElement = item
            if str(self.bot_data.get("AGENCIA")).lower() in item.text.lower():
                item.click()
                break      
     
    def proc_nattribut(self) -> None:
        
        numprocess = self.bot_data.get("NUMERO_PROCESSO").split(".")
        numproc_formated = f"{numprocess[0]}.{numprocess[1]}.{numprocess[3]}.{numprocess[4]}"

        self.interact.wait_caixa()
        self.message = "Informando numero do processo"
        self.type_log = "log"
        self.prt(self)()
        num_process: WebElement = self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:nuProcessoCNJ"]')))
        num_process.send_keys(numproc_formated)

        self.interact.wait_caixa()
        self.message = "Informando tipo da ação do processo"
        self.type_log = "log"
        self.prt(self)()
        list_type_acao_process = self.driver.find_element(By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:idOrigemAcao"]').find_elements(By.TAG_NAME, 'option')
        for item in list_type_acao_process:
            item: WebElement = item
            if str(self.bot_data.get("TIPO_ACAO")).lower() in item.text.lower():
                item.click()
                break

        self.interact.wait_caixa() 
        self.message = "Informando natureza tributaria"
        self.type_log = "log"
        self.prt(self)()
        natureza_tributaria = self.driver.find_element(By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:naturezaAcao"]').find_elements(By.TAG_NAME, 'option')[2]
        natureza_tributaria.click()
    
    def dados_partes(self) -> None:
            
        self.interact.wait_caixa()
        self.message = "Informando nome do autor"
        self.type_log = "log"
        self.prt(self)()
        campo_nome_autor = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:nomeAutor"]')
        campo_nome_autor.send_keys(self.bot_data.get("AUTOR"))
        
        
        self.interact.wait_caixa()
        self.message = "Informando tipo de documento do autor"
        self.type_log = "log"
        self.prt(self)()
        doct_type = count_doc(self.bot_data.get("CPF_CNPJ_AUTOR"))
        
        if not doct_type:
            return
        
        doctype_autor = self.driver.find_element(By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:tipoDocAutor"]').find_elements(By.TAG_NAME, 'option')
        
        for item in doctype_autor:
            item: WebElement = item
            if item.text.lower() == doct_type.lower():
                item.click()
                break
            
        self.interact.wait_caixa()
        self.message = "Informando documento do autor"
        self.type_log = "log"
        self.prt(self)()
        
        self.interact.wait_caixa()
        campo_doc_autor = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:codDocAutor"]')
        doc_autor = str(self.bot_data.get("CPF_CNPJ_AUTOR")).replace("-", "").replace(".", "").replace("/", "")
        campo_doc_autor.send_keys(doc_autor)

        self.interact.wait_caixa()
        meesage = "Informando réu"
        self.type_log = "log"
        self.prt(self)()
        campo_nome_reu = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:nomeReu"]')
        
        contraria = None
        for passivo in ["réu", "reu"]:
            contraria = self.bot_data.get(passivo.upper(), None)
            if contraria is not None:
                break
        
        campo_nome_reu.send_keys(self.bot_data.get("REU", contraria))
        
        doct_type = count_doc(self.bot_data.get("CPF_CNPJ_REU"))
        
        self.interact.wait_caixa()
        doctype_reu = self.driver.find_element(By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:tipoDocReu"]').find_elements(By.TAG_NAME, 'option')
        for item in doctype_reu:
            if item.text.lower() == doct_type.lower():
                item.click()
                break
        

        self.interact.wait_caixa()
        self.message = "Informando tipo de documento réu"
        self.type_log = "log"
        self.prt(self)()
        campo_doc_reu = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:codDocReu"]')
        doc_reu = str(self.bot_data.get("CPF_CNPJ_REU")).replace(".","").replace("-","").replace("/","")
        campo_doc_reu.send_keys(doc_reu)
            
    def info_deposito(self) -> None:
        
        self.interact.wait_caixa()
        self.message = "Informando indicador depositante"
        self.type_log = "log"
        self.prt(self)()
        indicador_depositante = self.driver.find_element(By.CSS_SELECTOR, 'select[id="j_id5:filtroView:formFormulario:idDepositante"]').find_elements(By.TAG_NAME, 'option')
        
        for item in indicador_depositante:
            
            if item.text.lower() == 'réu':
                item.click()
                break

        self.interact.wait_caixa()
        self.message = "Informando valor do depósito"
        self.type_log = "log"
        self.prt(self)()
        campo_val_deposito = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:valorDeposito"]')
        
        val_deposito = str(self.bot_data.get("VALOR_CALCULADO"))
        
        if not "," in val_deposito:
            val_deposito = f"{val_deposito},00"
        campo_val_deposito.send_keys(val_deposito)
    
    def make_doc(self) -> None:
    
        self.interact.wait_caixa()
        self.message = "Gerando documento"
        self.type_log = "log"
        self.prt(self)()
        make_id = self.driver.find_element(By.CSS_SELECTOR, 'input[id="j_id5:filtroView:formFormulario:j_id248"]')
        make_id.click()

        self.interact.wait_caixa()
        self.message = "Baixando documento"
        self.type_log = "log"
        self.prt(self)()
        download_pdf = self.driver.find_element(By.CSS_SELECTOR, 'a[id="j_id5:filtroView:formFormulario:j_id554"]')
        download_pdf.click()

    def rename_pdf(self) -> str:
        
        pgto_name = self.bot_data.get("NOME_CUSTOM", "Guia De Depósito")
        
        pdf_name = f"{pgto_name} - {self.bot_data.get('NUMERO_PROCESSO')} - {self.bot_data.get('AUTOR')} - {self.pid}.pdf"
        sleep(3)
        
        caminho_old_pdf = os.path.join(self.output_dir_path, "guia_boleto.pdf")
        renamepdf = os.path.join(self.output_dir_path, pdf_name)

        sleep(1)
        shutil.move(caminho_old_pdf, renamepdf)
        
        return pdf_name
                  
    def get_val_doc_and_codebar(self, pdf_name: str) -> None:
        
        sleep(0.5)
        
        path_pdf = os.path.join(self.output_dir_path, pdf_name)
        # Inicialize uma lista para armazenar os números encontrados
        bar_code = ''
        numeros_encontrados = []

        # Expressão regular para encontrar números nesse formato
        pattern = r'\b\d{5}\.\d{5}\s*\d{5}\.\d{6}\s*\d{5}\.\d{6}\s*\d\s*\d{14}\b'

        pdf_file = path_pdf
        read = PdfReader(pdf_file)

        for page in read.pages:
            text = page.extract_text()
            
            try:
                # Use a expressão regular para encontrar números
                numeros = re.findall(pattern, text)
                
                # Adicione os números encontrados à lista
                numeros_encontrados.extend(numeros)
            except:
                pass
            
        # Imprima os números encontrados
        for numero in numeros_encontrados:
            numero = str(numero)
            bar_code = numero.replace("  ","").replace(" ","").replace(".", " ") 

        return [self.bot_data.get("NUMERO_PROCESSO"), self.bot_data.get("TEXTO_DESC", ""), self.bot_data.get("VALOR_CALCULADO"), self.bot_data.get("DATA_PGTO", ""), 
                "condenação", "JEC", self.bot_data.get("VIA_CONDENACAO", ""), bar_code, pdf_name]
                